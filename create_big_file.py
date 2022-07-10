from openpyxl import Workbook
from random import randrange

def create_files(files):
    for file in files:
        wb = Workbook(write_only=True)
        ws = []
        ws.append(wb.create_sheet("Approval Or Provisioning"))
        ws.append(wb.create_sheet("Approval Or Provisioning (1)"))
        ws.append(wb.create_sheet("Approval Or Provisioning (2)"))
        ws.append(wb.create_sheet("Approval Or Provisioning (3)"))

        ws[0].append(['%d' % i for i in range(1, 35)])
        ws[1].append(['%d' % i for i in range(1, 35)])
        ws[2].append(['%d' % i for i in range(1, 35)])
        ws[3].append(['%d' % i for i in range(1, 35)])
        for row in range(2, 200):
            ws[0].append(['%d' % randrange(1,11) for i in range(1, 35)])
            ws[1].append(['%d' % randrange(1,11) for i in range(1, 35)])
            ws[2].append(['%d' % randrange(1,11) for i in range(1, 35)])
            ws[3].append(['%d' % randrange(1,11) for i in range(1, 35)])
        wb.save('dev/'+file)

if __name__ == "__main__":
    files = ['Approval_Or_Provisioning_juin.xlsx','Approval_Or_Provisioning_mai.xlsx']
    create_files(files)